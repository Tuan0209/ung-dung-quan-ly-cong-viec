import calendar as cal
import csv
import datetime
import json

from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User
from django.core.paginator import Paginator
from django.db.models import Case, Count, F, IntegerField, Q, When
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.http import require_POST

from .forms import (AttachmentForm, CategoryForm, CommentForm, SubTaskForm,
                    TagForm, TaskForm)
from .models import Attachment, Category, Comment, SubTask, Tag, Task


def _is_admin(user):
    return user.is_authenticated and user.is_staff


def _visible_tasks(user):
    """Công việc người dùng thấy: do mình tạo hoặc được giao, chưa ở thùng rác."""
    return Task.objects.filter(
        Q(owner=user) | Q(assigned_to=user), is_deleted=False
    ).distinct()


# ----------------------------- Bảng điều khiển ----------------------------- #


@login_required
def dashboard(request):
    tasks = _visible_tasks(request.user)
    today = timezone.localdate()

    total = tasks.count()
    done = tasks.filter(status=Task.Status.DONE).count()
    doing = tasks.filter(status=Task.Status.DOING).count()
    todo = tasks.filter(status=Task.Status.TODO).count()
    overdue = tasks.filter(
        due_date__lt=today, status__in=[Task.Status.TODO, Task.Status.DOING]
    ).count()
    completion_rate = round(done / total * 100) if total else 0

    priority_data = [
        tasks.filter(priority=Task.Priority.HIGH).count(),
        tasks.filter(priority=Task.Priority.MEDIUM).count(),
        tasks.filter(priority=Task.Priority.LOW).count(),
    ]

    # Hoàn thành theo 7 ngày gần nhất
    labels7, done7 = [], []
    for i in range(6, -1, -1):
        d = today - datetime.timedelta(days=i)
        labels7.append(d.strftime('%d/%m'))
        done7.append(tasks.filter(completed_at__date=d).count())

    upcoming = tasks.filter(
        due_date__gte=today,
        due_date__lte=today + datetime.timedelta(days=7),
        status__in=[Task.Status.TODO, Task.Status.DOING],
    ).order_by('due_date')[:5]
    recent = tasks.order_by('-created_at')[:5]

    context = {
        'total': total, 'done': done, 'doing': doing, 'todo': todo,
        'overdue': overdue, 'completion_rate': completion_rate,
        'starred_count': tasks.filter(is_starred=True).count(),
        'status_data': json.dumps([todo, doing, done]),
        'priority_data': json.dumps(priority_data),
        'labels7': json.dumps(labels7), 'done7': json.dumps(done7),
        'upcoming': upcoming, 'recent': recent,
    }
    return render(request, 'tasks/dashboard.html', context)


# ------------------------------- Công việc -------------------------------- #


@login_required
def task_list(request):
    tasks = _visible_tasks(request.user).select_related('category', 'assigned_to').prefetch_related('tags', 'subtasks')

    keyword = request.GET.get('q', '').strip()
    status = request.GET.get('status', '')
    priority = request.GET.get('priority', '')
    category_id = request.GET.get('category', '')
    tag_id = request.GET.get('tag', '')
    starred = request.GET.get('starred', '')
    sort = request.GET.get('sort', 'newest')

    if keyword:
        tasks = tasks.filter(Q(title__icontains=keyword) | Q(description__icontains=keyword))
    if status:
        tasks = tasks.filter(status=status)
    if priority:
        tasks = tasks.filter(priority=priority)
    if category_id:
        tasks = tasks.filter(category_id=category_id)
    if tag_id:
        tasks = tasks.filter(tags__id=tag_id)
    if starred == '1':
        tasks = tasks.filter(is_starred=True)

    if sort == 'oldest':
        tasks = tasks.order_by('created_at')
    elif sort == 'due':
        tasks = tasks.order_by(F('due_date').asc(nulls_last=True))
    elif sort == 'priority':
        order = Case(
            When(priority=Task.Priority.HIGH, then=0),
            When(priority=Task.Priority.MEDIUM, then=1),
            When(priority=Task.Priority.LOW, then=2),
            output_field=IntegerField(),
        )
        tasks = tasks.annotate(prio_order=order).order_by('prio_order', '-created_at')
    elif sort == 'title':
        tasks = tasks.order_by('title')
    else:
        sort = 'newest'
        tasks = tasks.order_by('-is_starred', '-created_at')

    paginator = Paginator(tasks, 8)
    page_obj = paginator.get_page(request.GET.get('page'))

    context = {
        'page_obj': page_obj,
        'categories': Category.objects.filter(owner=request.user),
        'tags': Tag.objects.filter(owner=request.user),
        'status_choices': Task.Status.choices,
        'priority_choices': Task.Priority.choices,
        'sort_choices': [
            ('newest', 'Mới nhất'), ('oldest', 'Cũ nhất'), ('due', 'Hạn gần nhất'),
            ('priority', 'Ưu tiên cao nhất'), ('title', 'Tên A → Z'),
        ],
        'current': {
            'q': keyword, 'status': status, 'priority': priority,
            'category': category_id, 'tag': tag_id, 'starred': starred, 'sort': sort,
        },
    }
    return render(request, 'tasks/task_list.html', context)


@login_required
def task_detail(request, pk):
    task = get_object_or_404(
        Task.objects.filter(Q(owner=request.user) | Q(assigned_to=request.user)), pk=pk
    )
    context = {
        'task': task,
        'subtask_form': SubTaskForm(),
        'comment_form': CommentForm(),
        'attachment_form': AttachmentForm(),
    }
    return render(request, 'tasks/task_detail.html', context)


@login_required
def task_create(request):
    if request.method == 'POST':
        form = TaskForm(request.POST, user=request.user)
        if form.is_valid():
            task = form.save(commit=False)
            task.owner = request.user
            task.save()
            form.save_m2m()
            messages.success(request, 'Đã thêm công việc mới.')
            return redirect('tasks:task_detail', pk=task.pk)
        messages.error(request, 'Vui lòng kiểm tra lại thông tin.')
    else:
        form = TaskForm(user=request.user)
    return render(request, 'tasks/task_form.html', {'form': form, 'title': 'Thêm công việc'})


@login_required
def task_update(request, pk):
    task = get_object_or_404(Task, pk=pk, owner=request.user)
    if request.method == 'POST':
        form = TaskForm(request.POST, instance=task, user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Đã cập nhật công việc.')
            return redirect('tasks:task_detail', pk=task.pk)
        messages.error(request, 'Vui lòng kiểm tra lại thông tin.')
    else:
        form = TaskForm(instance=task, user=request.user)
    return render(request, 'tasks/task_form.html',
                  {'form': form, 'title': 'Sửa công việc', 'task': task})


@login_required
@require_POST
def task_delete(request, pk):
    """Xóa mềm -> đưa vào thùng rác."""
    task = get_object_or_404(Task, pk=pk, owner=request.user)
    task.soft_delete()
    messages.success(request, 'Đã chuyển công việc vào thùng rác.')
    return redirect('tasks:task_list')


@login_required
@require_POST
def task_toggle(request, pk):
    task = get_object_or_404(
        Task.objects.filter(Q(owner=request.user) | Q(assigned_to=request.user)), pk=pk
    )
    task.status = Task.Status.TODO if task.status == Task.Status.DONE else Task.Status.DONE
    task.save()
    return redirect(request.META.get('HTTP_REFERER', 'tasks:task_list'))


@login_required
@require_POST
def task_star(request, pk):
    task = get_object_or_404(
        Task.objects.filter(Q(owner=request.user) | Q(assigned_to=request.user)), pk=pk
    )
    task.is_starred = not task.is_starred
    task.save(update_fields=['is_starred'])
    return redirect(request.META.get('HTTP_REFERER', 'tasks:task_list'))


# --------------------------- Công việc con -------------------------------- #


@login_required
@require_POST
def subtask_add(request, pk):
    task = get_object_or_404(Task, pk=pk, owner=request.user)
    form = SubTaskForm(request.POST)
    if form.is_valid():
        sub = form.save(commit=False)
        sub.task = task
        sub.save()
    else:
        messages.error(request, 'Nội dung công việc con không hợp lệ.')
    return redirect('tasks:task_detail', pk=pk)


@login_required
@require_POST
def subtask_toggle(request, pk):
    sub = get_object_or_404(SubTask, pk=pk, task__owner=request.user)
    sub.is_done = not sub.is_done
    sub.save(update_fields=['is_done'])
    return redirect('tasks:task_detail', pk=sub.task_id)


@login_required
@require_POST
def subtask_delete(request, pk):
    sub = get_object_or_404(SubTask, pk=pk, task__owner=request.user)
    task_id = sub.task_id
    sub.delete()
    return redirect('tasks:task_detail', pk=task_id)


# ------------------------------ Bình luận --------------------------------- #


@login_required
@require_POST
def comment_add(request, pk):
    task = get_object_or_404(
        Task.objects.filter(Q(owner=request.user) | Q(assigned_to=request.user)), pk=pk
    )
    form = CommentForm(request.POST)
    if form.is_valid():
        c = form.save(commit=False)
        c.task = task
        c.user = request.user
        c.save()
        messages.success(request, 'Đã thêm bình luận.')
    else:
        messages.error(request, 'Bình luận không hợp lệ.')
    return redirect('tasks:task_detail', pk=pk)


@login_required
@require_POST
def comment_delete(request, pk):
    comment = get_object_or_404(Comment, pk=pk, user=request.user)
    task_id = comment.task_id
    comment.delete()
    return redirect('tasks:task_detail', pk=task_id)


# ---------------------------- Tệp đính kèm -------------------------------- #


@login_required
@require_POST
def attachment_add(request, pk):
    task = get_object_or_404(
        Task.objects.filter(Q(owner=request.user) | Q(assigned_to=request.user)), pk=pk
    )
    form = AttachmentForm(request.POST, request.FILES)
    if form.is_valid():
        f = form.cleaned_data['file']
        Attachment.objects.create(
            task=task, file=f, original_name=f.name, uploaded_by=request.user
        )
        messages.success(request, 'Đã tải tệp đính kèm.')
    else:
        messages.error(request, 'Vui lòng chọn tệp hợp lệ.')
    return redirect('tasks:task_detail', pk=pk)


@login_required
@require_POST
def attachment_delete(request, pk):
    att = get_object_or_404(Attachment, pk=pk, task__owner=request.user)
    task_id = att.task_id
    att.file.delete(save=False)
    att.delete()
    return redirect('tasks:task_detail', pk=task_id)


# ------------------------------- Nhắc nhở --------------------------------- #


@login_required
@require_POST
def task_remind(request, pk):
    """Gửi email nhắc nhở về công việc cho người được giao."""
    from django.core.mail import send_mail
    from django.conf import settings

    task = get_object_or_404(
        Task.objects.filter(Q(owner=request.user) | Q(assigned_to=request.user)), pk=pk
    )
    target = task.assigned_to or task.owner
    if not target.email:
        messages.warning(request, 'Người nhận chưa có email để nhắc.')
        return redirect('tasks:task_detail', pk=pk)
    try:
        send_mail(
            subject=f'[Nhắc việc] {task.title}',
            message=(
                f'Xin chào {target.get_full_name() or target.username},\n\n'
                f'Bạn có công việc: "{task.title}"\n'
                f'Hạn hoàn thành: {task.due_date or "Chưa đặt"}\n'
                f'Trạng thái: {task.get_status_display()}\n\n'
                f'Vui lòng kiểm tra và hoàn thành đúng hạn.'
            ),
            from_email=getattr(settings, 'DEFAULT_FROM_EMAIL', 'noreply@congviec.vn'),
            recipient_list=[target.email],
            fail_silently=False,
        )
        messages.success(request, f'Đã gửi email nhắc nhở tới {target.email}.')
    except Exception as e:
        messages.error(request, f'Không gửi được email: {e}')
    return redirect('tasks:task_detail', pk=pk)


# ----------------------------- Bảng Kanban -------------------------------- #


@login_required
def kanban(request):
    tasks = _visible_tasks(request.user).select_related('category')
    cols = [
        (Task.Status.TODO, 'Cần làm', list(tasks.filter(status=Task.Status.TODO)), 'secondary'),
        (Task.Status.DOING, 'Đang làm', list(tasks.filter(status=Task.Status.DOING)), 'info'),
        (Task.Status.DONE, 'Hoàn thành', list(tasks.filter(status=Task.Status.DONE)), 'success'),
    ]
    return render(request, 'tasks/kanban.html', {'cols': cols})


@login_required
@require_POST
def task_set_status(request, pk):
    """Cập nhật trạng thái qua kéo-thả Kanban (AJAX)."""
    task = get_object_or_404(
        Task.objects.filter(Q(owner=request.user) | Q(assigned_to=request.user)), pk=pk
    )
    new_status = request.POST.get('status')
    if new_status in dict(Task.Status.choices):
        task.status = new_status
        task.save()
        return JsonResponse({'ok': True, 'status': task.get_status_display()})
    return JsonResponse({'ok': False}, status=400)


# ------------------------------- Lịch ------------------------------------- #


@login_required
def calendar_view(request):
    today = timezone.localdate()
    try:
        year = int(request.GET.get('year', today.year))
        month = int(request.GET.get('month', today.month))
    except (TypeError, ValueError):
        year, month = today.year, today.month
    if month < 1:
        month, year = 12, year - 1
    elif month > 12:
        month, year = 1, year + 1

    tasks = _visible_tasks(request.user).filter(
        due_date__year=year, due_date__month=month
    )
    by_day = {}
    for t in tasks:
        by_day.setdefault(t.due_date.day, []).append(t)

    cal.setfirstweekday(cal.MONDAY)
    weeks = cal.monthcalendar(year, month)
    grid = []
    for week in weeks:
        row = []
        for day in week:
            row.append({
                'day': day,
                'tasks': by_day.get(day, []) if day else [],
                'is_today': day == today.day and month == today.month and year == today.year,
            })
        grid.append(row)

    month_names = ['', 'Tháng 1', 'Tháng 2', 'Tháng 3', 'Tháng 4', 'Tháng 5',
                   'Tháng 6', 'Tháng 7', 'Tháng 8', 'Tháng 9', 'Tháng 10',
                   'Tháng 11', 'Tháng 12']
    context = {
        'grid': grid, 'year': year, 'month': month,
        'month_label': f'{month_names[month]} năm {year}',
        'prev_month': month - 1, 'prev_year': year,
        'next_month': month + 1, 'next_year': year,
        'weekday_labels': ['T2', 'T3', 'T4', 'T5', 'T6', 'T7', 'CN'],
    }
    return render(request, 'tasks/calendar.html', context)


# ------------------------------- Thùng rác -------------------------------- #


@login_required
def trash(request):
    tasks = Task.objects.filter(owner=request.user, is_deleted=True).order_by('-deleted_at')
    return render(request, 'tasks/trash.html', {'tasks': tasks})


@login_required
@require_POST
def task_restore(request, pk):
    task = get_object_or_404(Task, pk=pk, owner=request.user, is_deleted=True)
    task.restore()
    messages.success(request, 'Đã khôi phục công việc.')
    return redirect('tasks:trash')


@login_required
@require_POST
def task_purge(request, pk):
    """Xóa vĩnh viễn."""
    task = get_object_or_404(Task, pk=pk, owner=request.user, is_deleted=True)
    task.delete()
    messages.success(request, 'Đã xóa vĩnh viễn công việc.')
    return redirect('tasks:trash')


@login_required
@require_POST
def trash_empty(request):
    Task.objects.filter(owner=request.user, is_deleted=True).delete()
    messages.success(request, 'Đã dọn sạch thùng rác.')
    return redirect('tasks:trash')


# --------------------------- Xuất dữ liệu / In ---------------------------- #


def _export_queryset(request):
    return _visible_tasks(request.user).select_related('category').order_by('-created_at')


@login_required
def export_csv(request):
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = 'attachment; filename="cong_viec.csv"'
    response.write('﻿')  # BOM để Excel hiển thị tiếng Việt đúng
    writer = csv.writer(response)
    writer.writerow(['STT', 'Tiêu đề', 'Danh mục', 'Độ ưu tiên', 'Trạng thái',
                     'Hạn hoàn thành', 'Người được giao', 'Ngày tạo'])
    for i, t in enumerate(_export_queryset(request), 1):
        writer.writerow([
            i, t.title, t.category.name if t.category else '',
            t.get_priority_display(), t.get_status_display(),
            t.due_date.strftime('%d/%m/%Y') if t.due_date else '',
            (t.assigned_to.get_full_name() or t.assigned_to.username) if t.assigned_to else '',
            t.created_at.strftime('%d/%m/%Y %H:%M'),
        ])
    return response


@login_required
def export_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = 'Công việc'
    headers = ['STT', 'Tiêu đề', 'Danh mục', 'Độ ưu tiên', 'Trạng thái',
               'Hạn hoàn thành', 'Người được giao', 'Ngày tạo']
    ws.append(headers)
    header_fill = PatternFill('solid', fgColor='4F46E5')
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
    for i, t in enumerate(_export_queryset(request), 1):
        ws.append([
            i, t.title, t.category.name if t.category else '',
            t.get_priority_display(), t.get_status_display(),
            t.due_date.strftime('%d/%m/%Y') if t.due_date else '',
            (t.assigned_to.get_full_name() or t.assigned_to.username) if t.assigned_to else '',
            t.created_at.strftime('%d/%m/%Y %H:%M'),
        ])
    widths = [6, 40, 18, 14, 14, 16, 22, 18]
    for idx, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + idx)].width = w

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="cong_viec.xlsx"'
    wb.save(response)
    return response


@login_required
def print_view(request):
    tasks = _export_queryset(request)
    return render(request, 'tasks/print.html', {'tasks': tasks, 'now': timezone.now()})


# -------------------------------- Danh mục -------------------------------- #


@login_required
def category_list(request):
    categories = (
        Category.objects.filter(owner=request.user)
        .annotate(num_tasks=Count('tasks', filter=Q(tasks__is_deleted=False)))
        .order_by('name')
    )
    if request.method == 'POST':
        form = CategoryForm(request.POST, user=request.user)
        if form.is_valid():
            category = form.save(commit=False)
            category.owner = request.user
            category.save()
            messages.success(request, 'Đã thêm danh mục.')
            return redirect('tasks:category_list')
        messages.error(request, 'Vui lòng kiểm tra lại thông tin.')
    else:
        form = CategoryForm(user=request.user)
    return render(request, 'tasks/category_list.html', {'categories': categories, 'form': form})


@login_required
def category_update(request, pk):
    category = get_object_or_404(Category, pk=pk, owner=request.user)
    if request.method == 'POST':
        form = CategoryForm(request.POST, instance=category, user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Đã cập nhật danh mục.')
            return redirect('tasks:category_list')
        messages.error(request, 'Vui lòng kiểm tra lại thông tin.')
    else:
        form = CategoryForm(instance=category, user=request.user)
    return render(request, 'tasks/category_form.html', {'form': form, 'category': category})


@login_required
@require_POST
def category_delete(request, pk):
    category = get_object_or_404(Category, pk=pk, owner=request.user)
    category.delete()
    messages.success(request, 'Đã xóa danh mục.')
    return redirect('tasks:category_list')


# --------------------------------- Nhãn ----------------------------------- #


@login_required
def tag_list(request):
    tags = (
        Tag.objects.filter(owner=request.user)
        .annotate(num_tasks=Count('tasks', filter=Q(tasks__is_deleted=False)))
        .order_by('name')
    )
    if request.method == 'POST':
        form = TagForm(request.POST, user=request.user)
        if form.is_valid():
            tag = form.save(commit=False)
            tag.owner = request.user
            tag.save()
            messages.success(request, 'Đã thêm nhãn.')
            return redirect('tasks:tag_list')
        messages.error(request, 'Vui lòng kiểm tra lại thông tin.')
    else:
        form = TagForm(user=request.user)
    return render(request, 'tasks/tag_list.html', {'tags': tags, 'form': form})


@login_required
@require_POST
def tag_delete(request, pk):
    tag = get_object_or_404(Tag, pk=pk, owner=request.user)
    tag.delete()
    messages.success(request, 'Đã xóa nhãn.')
    return redirect('tasks:tag_list')


# --------------------------- Quản trị hệ thống ---------------------------- #


@login_required
@user_passes_test(_is_admin)
def admin_dashboard(request):
    today = timezone.localdate()
    all_tasks = Task.objects.filter(is_deleted=False)

    # Thống kê theo trạng thái toàn hệ thống
    status_data = [
        all_tasks.filter(status=Task.Status.TODO).count(),
        all_tasks.filter(status=Task.Status.DOING).count(),
        all_tasks.filter(status=Task.Status.DONE).count(),
    ]
    # Top người dùng theo số công việc
    top_users = (
        User.objects.annotate(num_tasks=Count('tasks', filter=Q(tasks__is_deleted=False)))
        .order_by('-num_tasks')[:5]
    )
    context = {
        'total_users': User.objects.count(),
        'active_users': User.objects.filter(is_active=True).count(),
        'total_tasks': all_tasks.count(),
        'total_done': all_tasks.filter(status=Task.Status.DONE).count(),
        'total_overdue': all_tasks.filter(
            due_date__lt=today, status__in=[Task.Status.TODO, Task.Status.DOING]
        ).count(),
        'total_categories': Category.objects.count(),
        'total_tags': Tag.objects.count(),
        'trashed': Task.objects.filter(is_deleted=True).count(),
        'status_data': json.dumps(status_data),
        'top_users': top_users,
        'new_users': User.objects.order_by('-date_joined')[:5],
        'recent_tasks': all_tasks.select_related('owner', 'category').order_by('-created_at')[:8],
    }
    return render(request, 'tasks/admin_dashboard.html', context)


@login_required
@user_passes_test(_is_admin)
def admin_users(request):
    keyword = request.GET.get('q', '').strip()
    users = User.objects.annotate(
        num_tasks=Count('tasks', filter=Q(tasks__is_deleted=False))
    ).order_by('-date_joined')
    if keyword:
        users = users.filter(
            Q(username__icontains=keyword) | Q(first_name__icontains=keyword) | Q(email__icontains=keyword)
        )
    paginator = Paginator(users, 12)
    page_obj = paginator.get_page(request.GET.get('page'))
    return render(request, 'tasks/admin_users.html', {'page_obj': page_obj, 'q': keyword})


@login_required
@user_passes_test(_is_admin)
@require_POST
def admin_toggle_user(request, pk):
    target = get_object_or_404(User, pk=pk)
    if target == request.user:
        messages.error(request, 'Bạn không thể tự khóa tài khoản của mình.')
    else:
        target.is_active = not target.is_active
        target.save()
        state = 'mở khóa' if target.is_active else 'khóa'
        messages.success(request, f'Đã {state} tài khoản {target.username}.')
    return redirect(request.META.get('HTTP_REFERER', 'tasks:admin_users'))


@login_required
@user_passes_test(_is_admin)
@require_POST
def admin_toggle_staff(request, pk):
    """Cấp / thu hồi quyền quản trị viên."""
    target = get_object_or_404(User, pk=pk)
    if target == request.user:
        messages.error(request, 'Bạn không thể tự thay đổi quyền của mình.')
    else:
        target.is_staff = not target.is_staff
        target.save()
        state = 'cấp' if target.is_staff else 'thu hồi'
        messages.success(request, f'Đã {state} quyền quản trị cho {target.username}.')
    return redirect(request.META.get('HTTP_REFERER', 'tasks:admin_users'))


@login_required
@user_passes_test(_is_admin)
@require_POST
def admin_delete_user(request, pk):
    target = get_object_or_404(User, pk=pk)
    if target == request.user:
        messages.error(request, 'Bạn không thể tự xóa tài khoản của mình.')
    elif target.is_superuser:
        messages.error(request, 'Không thể xóa tài khoản quản trị tối cao.')
    else:
        name = target.username
        target.delete()
        messages.success(request, f'Đã xóa tài khoản {name}.')
    return redirect('tasks:admin_users')


@login_required
@user_passes_test(_is_admin)
def admin_tasks(request):
    """Quản lý toàn bộ công việc của mọi người dùng."""
    tasks = Task.objects.filter(is_deleted=False).select_related('owner', 'category')
    keyword = request.GET.get('q', '').strip()
    status = request.GET.get('status', '')
    owner_id = request.GET.get('owner', '')
    if keyword:
        tasks = tasks.filter(title__icontains=keyword)
    if status:
        tasks = tasks.filter(status=status)
    if owner_id:
        tasks = tasks.filter(owner_id=owner_id)
    tasks = tasks.order_by('-created_at')
    paginator = Paginator(tasks, 15)
    page_obj = paginator.get_page(request.GET.get('page'))
    context = {
        'page_obj': page_obj,
        'users': User.objects.order_by('username'),
        'status_choices': Task.Status.choices,
        'current': {'q': keyword, 'status': status, 'owner': owner_id},
    }
    return render(request, 'tasks/admin_tasks.html', context)


@login_required
@user_passes_test(_is_admin)
@require_POST
def admin_task_delete(request, pk):
    task = get_object_or_404(Task, pk=pk)
    task.delete()
    messages.success(request, 'Đã xóa công việc khỏi hệ thống.')
    return redirect(request.META.get('HTTP_REFERER', 'tasks:admin_tasks'))
